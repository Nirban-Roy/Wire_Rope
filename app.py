import os
import json
import io
import time
import threading
import sqlite3
from datetime import datetime
from threading import Lock

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import paho.mqtt.client as mqtt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# --- CONFIGURATION ---------------------------------------------------------

MQTT_BROKER   = "3.230.164.113"
MQTT_PORT     = 1883
MQTT_USER     = "IEMA@2024"
MQTT_PASSWORD = "Pass@IEMA2024"
DEVICES_DB    = "devices.db"

# Header styling for Excel reports
HEADER_FILL  = PatternFill(fill_type="solid", fgColor="FFD966")  # light gold
HEADER_FONT  = Font(bold=True, color="000000")
ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")  # light grey

# --- IN-MEMORY STATE ------------------------------------------------------

data_lock = Lock()

sensor_data = {
    'sensors': {i: 0.0 for i in range(1, 9)},
    'imu':     {'accelX': 0.0, 'accelY': 0.0, 'accelZ': 0.0,
                'gyroX':  0.0, 'gyroY':  0.0, 'gyroZ':  0.0},
    'steps':   0,
    'distance':0.0,
}

log_rows = []
HEADERS = [
    'Timestamp', 'Rope Length (m)', 'Inspected Length (m)',
    'Anomaly Count', 'Steps Count', 'Distance (m)',
] + [f"Sensor {i} (Gauss)" for i in range(1,9)] + [
    'Accel X (m/s²)', 'Accel Y (m/s²)', 'Accel Z (m/s²)',
    'Gyro X (rad/s)', 'Gyro Y (rad/s)', 'Gyro Z (rad/s)',
]
log_rows.append(HEADERS)

# --- SQLITE INITIALIZATION ------------------------------------------------

def init_db():
    conn = sqlite3.connect(DEVICES_DB)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS devices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_name TEXT NOT NULL,
            device_id   TEXT NOT NULL UNIQUE,
            mqtt_topics TEXT NOT NULL,
            created_at  TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

try:
    init_db()
    print("✅ Database initialized successfully.")
except Exception as e:
    print("❌ Failed to initialize database:", e)


# --- MQTT CALLBACKS -------------------------------------------------------

def on_message(client, userdata, msg):
    topic = msg.topic
    payload = msg.payload.decode('utf-8')
    with data_lock:
        try:
            if topic.startswith("wire_rope_"):
                # parse the trailing number
                idx = int(topic.rsplit('_', 1)[1])
                if 1 <= idx <= 8:
                    # Hall‐effect sensors 1–8
                    sensor_data['sensors'][idx] = float(payload)
                elif idx == 9:
                    # IMU topic now contains pure JSON
                    imu = json.loads(payload)
                    # update accelX/Y/Z
                    for axis in ('X','Y','Z'):
                        sensor_data['imu'][f"accel{axis}"] = imu.get(f"accel{axis}", 0.0)
                    # update gyroX/Y/Z
                    for axis in ('X','Y','Z'):
                        sensor_data['imu'][f"gyro{axis}"]  = imu.get(f"gyro{axis}",  0.0)

            elif topic == "steps":
                sensor_data['steps'] = int(payload)

            elif topic == "distance":
                sensor_data['distance'] = float(payload)

            # ignore everything else
        except Exception as e:
            print("❌ Error processing MQTT message:", e)

def on_connect(client, userdata, flags, rc):
    print(f"✅ Connected to MQTT broker, rc={rc}")
mqtt_client = mqtt.Client()
mqtt_client.username_pw_set(MQTT_USER, MQTT_PASSWORD)
mqtt_client.on_connect = on_connect
mqtt_client.on_message = on_message
mqtt_client.connect(MQTT_BROKER, MQTT_PORT, keepalive=60)
mqtt_client.loop_start()

# --- BACKGROUND LOGGER ----------------------------------------------------

def snapshot_logger():
    while True:
        with data_lock:
            # take a snapshot of whatever is currently in sensor_data
            row = [
                datetime.now().isoformat(),
                0.0, 0.0, 0,                     # your rope length / inspected length / anomaly count
                sensor_data['steps'],
                sensor_data['distance'],
            ] + [sensor_data['sensors'][i] for i in range(1, 9)] + [
                sensor_data['imu']['accelX'],
                sensor_data['imu']['accelY'],
                sensor_data['imu']['accelZ'],
                sensor_data['imu']['gyroX'],
                sensor_data['imu']['gyroY'],
                sensor_data['imu']['gyroZ'],
            ]
            log_rows.append(row)
        time.sleep(1)


threading.Thread(target=snapshot_logger, daemon=True).start()

# --- FLASK APP ------------------------------------------------------------
app = Flask(__name__, static_folder="static")
CORS(app)

# ————————————————————————————————
# Dynamic subscription endpoint
# ————————————————————————————————

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    return render_template('landing.html')

@app.route('/admin')
def admin():
    conn = sqlite3.connect(DEVICES_DB)
    c = conn.cursor()
    c.execute("SELECT device_name, device_id, mqtt_topics, created_at FROM devices")
    rows = c.fetchall()
    conn.close()
    devices = [
        {'device_name': name,
         'device_id': did,
         'mqtt_topics': topics.split(','),
         'created_at': created}
        for name, did, topics, created in rows
    ]
    return render_template('admin.html', devices=devices)

# --- Device CRUD APIs -----------------------------------------------------

@app.route('/api/add-device', methods=['POST'])
def add_device():
    data = request.get_json(force=True)
    name, did, topics = data['device_name'], data['device_id'], data['mqtt_topics']
    created = datetime.now().isoformat()
    try:
        conn = sqlite3.connect(DEVICES_DB)
        c = conn.cursor()
        c.execute("""
            INSERT INTO devices (device_name, device_id, mqtt_topics, created_at)
            VALUES (?, ?, ?, ?)
        """, (name, did, ','.join(topics), created))
        conn.commit()
        conn.close()
        return jsonify({'status':'Device added'}), 200
    except sqlite3.IntegrityError as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/edit-device', methods=['POST'])
def edit_device():
    data = request.get_json(force=True)
    orig_id = data['original_device_id']
    name, new_id, topics = data['device_name'], data['device_id'], data['mqtt_topics']
    try:
        conn = sqlite3.connect(DEVICES_DB)
        c = conn.cursor()
        c.execute("""
            UPDATE devices
               SET device_name = ?, device_id = ?, mqtt_topics = ?
             WHERE device_id = ?
        """, (name, new_id, ','.join(topics), orig_id))
        conn.commit()
        conn.close()
        return jsonify({'status':'Device updated'}), 200
    except sqlite3.IntegrityError as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/delete-device', methods=['POST'])
def delete_device():
    data = request.get_json(force=True)
    did = data['device_id']
    conn = sqlite3.connect(DEVICES_DB)
    c = conn.cursor()
    c.execute("DELETE FROM devices WHERE device_id = ?", (did,))
    conn.commit()
    conn.close()
    return jsonify({'status':'Device deleted'}), 200

# --- Sensor Data & Report APIs --------------------------------------------
@app.route('/api/devices')
def get_devices():
    conn = sqlite3.connect(DEVICES_DB)
    c = conn.cursor()
    c.execute("SELECT device_name, device_id FROM devices")
    rows = c.fetchall()
    conn.close()
    return jsonify([
        {'name': name, 'id': did}
        for name, did in rows
    ])

@app.route('/api/sensor-data')
def get_sensor_data():
    with data_lock:
        return jsonify(sensor_data)

@app.route('/api/download-report', methods=['GET','POST'])
def download_report():
    wb = Workbook()
    ws = wb.active
    for col, title in enumerate(log_rows[0], start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for r, row in enumerate(log_rows[1:], start=2):
        fill = ALT_ROW_FILL if r % 2 == 0 else None
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if fill: cell.fill = fill
    stream = io.BytesIO()
    wb.save(stream); stream.seek(0)
    fname = f"wire_rope_full_log_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    return send_file(
        stream, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# near top of file, under your imports and after mqtt_client.loop_start():
last_topics = []

@app.route('/api/select-device')
def select_device():
    global last_topics
    device_id = request.args.get('device_id')
    if not device_id:
        return jsonify({'error': 'Device ID is required'}), 400

    # 1) pull raw comma-sep string
    conn = sqlite3.connect(DEVICES_DB)
    c = conn.cursor()
    c.execute("SELECT mqtt_topics FROM devices WHERE device_id = ?", (device_id,))
    row = c.fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Device not found'}), 404

    # 2) split, strip empty → pad to 11 slots with None
    raw = [t.strip() for t in row[0].split(',')]
    topics = raw + [None] * (11 - len(raw))
    topics = topics[:11]

    # 3) unsubscribe *each* old topic
    for t in last_topics:
        if t is not None:
            mqtt_client.unsubscribe(t)

    # 4) subscribe *only* non-None new topics
    new_subs = []
    for t in topics:
        if t:
            mqtt_client.subscribe(t)
            new_subs.append(t)

    last_topics = new_subs

    # 5) reset all in-memory data + log
    with data_lock:
        for i in range(1, 9):
            sensor_data['sensors'][i] = 0.0
        for k in sensor_data['imu']:
            sensor_data['imu'][k] = 0.0
        sensor_data['distance'] = 0.0
        sensor_data['steps']   = 0
        global log_rows
        log_rows = [HEADERS.copy()]

    # 6) report any truly missing slots
    missing = [f"slot#{i+1}" for i,t in enumerate(topics) if not t]
    resp = {'status': f'Subscribed to topics for {device_id}'}
    if missing:
        resp.update({
            'incomplete': True,
            'missing_topics': missing
        })
    return jsonify(resp), 200

@app.route('/api/unsubscribe-all', methods=['POST'])
def unsubscribe_all():
    mqtt_client.unsubscribe("#")
    return jsonify({'status':'Unsubscribed from all topics'}), 200

@app.route('/api/reset-report', methods=['POST'])
def reset_report():
    with data_lock:
        global log_rows
        log_rows = [HEADERS.copy()]
        sensor_data.update({'distance':0.0, 'steps':0})
        for i in sensor_data['sensors']: sensor_data['sensors'][i] = 0.0
        for axis in sensor_data['imu']: sensor_data['imu'][axis] = 0.0
    try:
        mqtt_client.reconnect()
    except:
        mqtt_client.disconnect()
        mqtt_client.connect(MQTT_BROKER, MQTT_PORT, keepalive=60)
        mqtt_client.loop_start()
    return jsonify({'status':'Report log reset'}), 200

@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_file(os.path.join("static", filename))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
